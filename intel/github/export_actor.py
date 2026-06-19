#!/usr/bin/env python3
"""Export standardgalactic counter-intel network to a dedicated xlsx + append to main inventory."""
import os, sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

WS = os.path.expanduser("~/dev/gh-inventory")
db = sqlite3.connect(os.path.join(WS, "inventory.db"))
db.row_factory = sqlite3.Row

wb = Workbook()

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy for SG report
SEC_FILL     = PatternFill("solid", fgColor="C00000")   # red highlight for sec-flagged
NOTABLE_FILL = PatternFill("solid", fgColor="FF6600")   # orange for notable

def add_sheet(title, headers, rows, widths, flag_col=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for ri, row in enumerate(rows, 2):
        ws.append(list(row))
        if flag_col:
            val = ws.cell(ri, flag_col).value
            if val == 'SEC':
                for c in range(1, len(headers)+1):
                    ws.cell(ri, c).fill = SEC_FILL
    ws.freeze_panes = "A2"
    if rows: ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def q(sql, p=()): return db.execute(sql, p).fetchall()

# ---- 1. SG Overview ----
total    = db.execute("SELECT count(*) FROM sg_accounts").fetchone()[0]
sec      = db.execute("SELECT count(*) FROM sg_accounts WHERE security_signal=1").fetchone()[0]
notable  = db.execute("SELECT count(*) FROM sg_accounts WHERE notable=1").fetchone()[0]
sg_fol   = db.execute("SELECT count(*) FROM sg_accounts WHERE relation='sg_follower'").fetchone()[0]
sg_ing   = db.execute("SELECT count(*) FROM sg_accounts WHERE relation='sg_following'").fetchone()[0]
overlap  = db.execute("""SELECT count(*) FROM sg_accounts s
    JOIN accounts a ON a.login=s.login WHERE a.relation IN ('follower','following','mutual')""").fetchone()[0]
your_sg_overlap = db.execute("""SELECT count(*) FROM sg_accounts s
    JOIN accounts a ON a.login=s.login WHERE a.security_signal=1""").fetchone()[0]

ws = wb.active; ws.title = "SG_Overview"
overview = [
    ["TARGET", "standardgalactic"],
    ["GitHub", "https://github.com/standardgalactic"],
    ["Bio", "𝘏𝘰𝘮𝘦 𝘰𝘧 𝘵𝘩𝘦 Standard Galactic Alphabet"],
    ["Company", "Xanadu"],
    ["Location", "Canada"],
    ["Following", "1,122,007 (mass-follow / network mapper)"],
    ["Followers", "22,503"],
    ["",""],
    ["=== SG NETWORK CRAWL ===", ""],
    ["SG followers pulled", sg_fol],
    ["SG following sampled", sg_ing],
    ["Total SG-network accounts", total],
    ["Security-flagged in SG network", sec],
    ["Notable (500+ followers or sec)", notable],
    ["",""],
    ["=== OVERLAP WITH YOUR NETWORK ===", ""],
    ["SG accounts also in YOUR 1-hop", overlap],
    ["SG accounts security-flagged AND in your 1-hop", your_sg_overlap],
    ["",""],
    ["=== ASSESSMENT ===", ""],
    ["Follow pattern", "1.1M following is not human — automated network mapping or influence op"],
    ["Risk", "SG is documenting who is in the AI/security/Linux space at scale"],
    ["Action", "Review SG network overlaps with your accounts — flag shared notable contacts"],
]
for row in overview: ws.append(row)
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 70
ws.cell(1,1).font = Font(bold=True, color="FFFFFF")
ws.cell(1,1).fill = HEADER_FILL

# ---- 2. SG Network — Security Flagged (priority list) ----
rows = q("""SELECT s.login,
    CASE WHEN a.login IS NOT NULL THEN 'YES — '||a.relation ELSE '' END as in_your_net,
    s.relation, s.name, s.followers_n, s.following_n, s.public_repos,
    substr(s.bio,1,200), s.company, s.location, s.interests, s.url
    FROM sg_accounts s LEFT JOIN accounts a ON a.login=s.login
    WHERE s.security_signal=1
    ORDER BY (a.login IS NOT NULL) DESC, s.followers_n DESC""")
add_sheet("SG_Security_Flagged",
    ["login","in_YOUR_network","sg_relation","name","followers","following","repos",
     "bio","company","location","interests","url"],
    rows, [24,20,14,24,9,9,7,60,24,20,30,42])

# ---- 3. SG Network — Notable (high-follower) ----
rows = q("""SELECT s.login,
    CASE WHEN a.login IS NOT NULL THEN 'YES — '||a.relation ELSE '' END,
    CASE WHEN s.security_signal THEN 'SEC' ELSE '' END,
    s.relation, s.name, s.followers_n, s.following_n,
    substr(s.bio,1,160), s.company, s.location, s.interests, s.url
    FROM sg_accounts s LEFT JOIN accounts a ON a.login=s.login
    WHERE s.notable=1
    ORDER BY s.followers_n DESC""")
add_sheet("SG_Notable",
    ["login","in_YOUR_network","sec","sg_relation","name","followers","following",
     "bio","company","location","interests","url"],
    rows, [24,20,5,14,24,9,9,50,24,20,28,42], flag_col=3)

# ---- 4. Overlap — accounts in BOTH your network AND SG's network ----
rows = q("""SELECT s.login, a.relation as your_rel, s.relation as sg_rel,
    CASE WHEN s.security_signal THEN 'SEC' ELSE '' END,
    CASE WHEN a.security_signal THEN 'SEC' ELSE '' END,
    s.name, s.followers_n, substr(s.bio,1,200), s.company, s.location, s.interests, s.url
    FROM sg_accounts s JOIN accounts a ON a.login=s.login
    ORDER BY (s.security_signal=1 OR a.security_signal=1) DESC, s.followers_n DESC""")
add_sheet("Overlap_YourNet_AND_SG",
    ["login","your_relation","sg_relation","sg_sec","your_sec","name","followers",
     "bio","company","location","interests","url"],
    rows, [24,12,14,8,8,24,9,60,24,20,30,42])

# ---- 5. Full SG account list ----
rows = q("""SELECT s.login, s.relation,
    CASE WHEN s.security_signal THEN 'SEC' ELSE '' END,
    CASE WHEN s.notable THEN 'notable' ELSE '' END,
    CASE WHEN a.login IS NOT NULL THEN a.relation ELSE '' END,
    s.name, s.followers_n, s.following_n, s.public_repos,
    substr(s.bio,1,160), s.company, s.location, s.interests,
    substr(s.created_at,1,10), s.url
    FROM sg_accounts s LEFT JOIN accounts a ON a.login=s.login
    ORDER BY s.security_signal DESC, s.followers_n DESC""")
add_sheet("SG_All_Accounts",
    ["login","sg_relation","sec","notable","in_your_net","name","followers","following",
     "repos","bio","company","location","interests","joined","url"],
    rows, [24,14,5,8,12,24,9,9,7,50,24,20,28,12,42], flag_col=3)

# ---- 6. SG edges ----
rows = q("SELECT src, dst FROM sg_edges ORDER BY src")
add_sheet("SG_Edges",
    ["src","dst"], rows, [30,30])

wb.move_sheet("SG_Overview", -(len(wb.sheetnames)-1))
del wb["Sheet"]

out = os.path.join(WS, "standardgalactic_intel.xlsx")
wb.save(out)
print("WROTE", out)
for s in wb.sheetnames:
    print(f"  - {s}: {wb[s].max_row-1} rows")
