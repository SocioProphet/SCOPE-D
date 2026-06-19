#!/usr/bin/env python3
"""Extract learning signals from the forks catalog.

Scores each fork by strategic value to the SocioProphet platform and writes
learn_notes back into the DB, then exports a prioritized xlsx report.

Scoring: category relevance × upstream activity × staleness penalty
Priority categories: Security > AI/ML > Platform > Linux > Cloud > DevTools/CLI
"""
import os, re, sqlite3
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

WS   = Path(os.path.expanduser("~/dev/gh-inventory"))
db   = sqlite3.connect(WS / "inventory.db")
db.row_factory = sqlite3.Row

CATEGORY_SCORE = {
    "Security":     10,
    "AI/ML":         9,
    "Platform":      8,
    "Linux":         7,
    "Cloud":         7,
    "DevTools/CLI":  6,
    "Data/ML-Ops":   5,
    "Web/Frontend":  3,
    "Other/Uncategorized": 1,
}

LESSON_PATTERNS = [
    (re.compile(r"mesh|service.mesh|istio|envoy", re.I),     "Service mesh architecture — routing, discovery, multi-cluster patterns"),
    (re.compile(r"rag|retrieval.augment|vector.store|embed", re.I), "RAG / vector retrieval architecture"),
    (re.compile(r"agent|agentic|tool.call|function.call", re.I),    "Agent / tool-use design pattern"),
    (re.compile(r"graph|knowledge.graph|neo4j|rdf|sparql", re.I),   "Graph data model — knowledge graph / ontology design"),
    (re.compile(r"osint|intel|threat|recon|shodan", re.I),           "OSINT / threat intelligence tooling"),
    (re.compile(r"mcp|model.context.protocol", re.I),                "MCP server architecture"),
    (re.compile(r"k8s|kubernetes|helm|operator", re.I),              "Kubernetes operator / Helm chart patterns"),
    (re.compile(r"wasm|webassembly", re.I),                          "WebAssembly runtime / module design"),
    (re.compile(r"p2p|peer.to.peer|libp2p|ipfs", re.I),             "P2P networking / distributed identity"),
    (re.compile(r"grpc|protobuf|proto3", re.I),                      "gRPC / protobuf schema-first API design"),
    (re.compile(r"lsp|language.server|tree.sitter|parser", re.I),   "Language server / AST / Tree-sitter — SynapseIQ relevant"),
    (re.compile(r"llm|gpt|claude|openai|anthropic|mistral|ollama", re.I), "LLM integration pattern"),
    (re.compile(r"federat|multicluster|multi.region", re.I),         "Federated / multi-region deployment pattern"),
    (re.compile(r"policy|rego|opa|policy.as.code", re.I),            "Policy-as-code / OPA / Rego — detections lane"),
    (re.compile(r"sigma|yara|snort|suricata|detection", re.I),       "Detection rules / SIEM — detections lane"),
    (re.compile(r"encrypt|kms|vault|secret|hsm", re.I),              "Secrets management / encryption — ai-infra hardening"),
    (re.compile(r"workflow|dag|airflow|prefect|luigi", re.I),        "Workflow orchestration DAG pattern"),
    (re.compile(r"rust|cargo|tokio|async.runtime", re.I),            "Rust async runtime — SourceOS / socios-linux relevant"),
    (re.compile(r"ebpf|bpf|kernel|syscall|seccomp", re.I),          "eBPF / kernel instrumentation — Linux security"),
    (re.compile(r"oauth|jwt|oidc|saml|sso|auth", re.I),             "Auth / identity pattern — boundary enforcement"),
    (re.compile(r"contract|schema|json.schema|openapi|asyncapi", re.I), "Contract-first / schema-first API — SCOPE-D pattern"),
    (re.compile(r"reinforcement|rl|reward|policy.gradient", re.I),   "Reinforcement learning — agent reward design"),
    (re.compile(r"static.analy|ast|lint|code.quality|semgrep", re.I),"Static analysis / AST tooling — codebase graph lane"),
    (re.compile(r"browser|webdriver|selenium|playwright|puppeteer", re.I), "Browser automation — BearBrowser relevant"),
    (re.compile(r"fingerprint|canary|honeypot|deception", re.I),     "Deception / fingerprinting — honeypot lane"),
    (re.compile(r"supply.chain|sbom|provenance|sigstore", re.I),     "Supply chain security / SBOM / provenance"),
    (re.compile(r"search|full.text|elastic|lucene|solr", re.I),      "Full-text search / indexing pattern"),
    (re.compile(r"terminal|tui|cli.framework|cobra|click", re.I),    "Terminal / TUI / CLI framework pattern"),
    (re.compile(r"distributed|consensus|raft|paxos", re.I),          "Distributed consensus protocol"),
    (re.compile(r"monitor|observ|tracing|otel|opentelemetry|prometheus", re.I), "Observability / telemetry — monitoring lane"),
]

def score_fork(row) -> tuple[int, str]:
    cat_str = row["category"] or "Other/Uncategorized"
    cats = [c.strip() for c in cat_str.split(";")]
    cat_score = max((CATEGORY_SCORE.get(c, 1) for c in cats), default=1)

    stars = row["upstream_stars"] or 0
    star_score = min(int((stars ** 0.4)), 15)

    staleness = row["staleness_years"] or 10
    stale_penalty = min(int(staleness * 1.5), 12)

    desc = (row["upstream_desc"] or "") + " " + (row["topics"] or "") + " " + (row["full_name"] or "")

    lessons = []
    for pat, note in LESSON_PATTERNS:
        if pat.search(desc):
            lessons.append(note)

    lesson_bonus = min(len(lessons) * 2, 8)
    total = cat_score + star_score - stale_penalty + lesson_bonus

    notes = "; ".join(lessons) if lessons else f"Fork in {cats[0]} category — review upstream for design patterns"
    if staleness > 5:
        notes += f" [STALE: {staleness:.1f}yr — upstream may be abandoned]"
    elif staleness < 0.5:
        notes += " [ACTIVE — upstream still maintained]"

    return total, notes


rows = db.execute("""
    SELECT full_name, bucket, upstream, upstream_desc, upstream_lang,
           upstream_stars, upstream_pushed, upstream_archived,
           staleness_years, upstream_age_years, category, language,
           topics, our_url, upstream_url
    FROM forks_catalog
""").fetchall()

scored = []
for r in rows:
    s, notes = score_fork(r)
    scored.append((s, notes, r))

scored.sort(key=lambda x: -x[0])

db.execute("BEGIN")
for s, notes, r in scored:
    db.execute("UPDATE forks_catalog SET learn_notes=? WHERE full_name=?", (notes, r["full_name"]))
db.execute("COMMIT")
print(f"Updated learn_notes for {len(scored)} forks")

wb = Workbook()
ws = wb.active
ws.title = "Fork_Learning"

HDR = PatternFill("solid", fgColor="1F3864")
HI  = PatternFill("solid", fgColor="375623")

headers = ["score","full_name","category","upstream","upstream_desc","lang",
           "stars","staleness_yr","learn_notes","upstream_url","our_url"]
ws.append(headers)
for c in range(1, len(headers)+1):
    cell = ws.cell(1,c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HDR

for s, notes, r in scored:
    row = [s, r["full_name"], r["category"], r["upstream"],
           (r["upstream_desc"] or "")[:120], r["upstream_lang"],
           r["upstream_stars"], r["staleness_years"], notes,
           r["upstream_url"], r["our_url"]]
    ws.append(row)
    if s >= 15:
        for c in range(1, len(row)+1):
            ws.cell(ws.max_row, c).fill = HI

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions
for i, w in enumerate([6,40,20,35,60,10,8,12,80,45,45], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out = WS / "fork_learning.xlsx"
wb.save(out)
print(f"Wrote {out}")

cats = defaultdict(list)
for s, notes, r in scored:
    for c in (r["category"] or "Other").split(";"):
        cats[c.strip()].append(s)

print("\nTop 30 forks by strategic score:")
for s, notes, r in scored[:30]:
    print(f"  [{s:3}] {r['full_name']:45} {(r['category'] or '')[:25]:25} stars={r['upstream_stars'] or 0:5}")

print(f"\nCategory summary (mean score / count):")
for cat in sorted(cats, key=lambda c: -sum(cats[c])/max(len(cats[c]),1)):
    scores = cats[cat]
    print(f"  {cat:25}  n={len(scores):4}  mean={sum(scores)/len(scores):.1f}  top={max(scores)}")

db.close()
