#!/usr/bin/env python3
"""Export the full inventory + counter-intel DB into a labeled multi-sheet xlsx.
Re-runnable: reflects whatever is currently in inventory.db."""
import os, sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WS = os.path.expanduser("~/dev/gh-inventory")
db = sqlite3.connect(os.path.join(WS, "inventory.db"))
db.row_factory = sqlite3.Row
ME = "mdheller"

wb = Workbook()
def add_sheet(title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(list(r))
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def q(sql, p=()): return db.execute(sql, p).fetchall()

# ---------- 1. Repos ----------
rows = q("""SELECT bucket,owner,name,role,category,
  CASE WHEN private THEN 'private' ELSE 'public' END vis,
  CASE WHEN is_fork THEN 'yes' ELSE '' END fork,
  CASE WHEN archived THEN 'yes' ELSE '' END arch,
  language,stars,forks,substr(pushed_at,1,10),substr(created_at,1,10),
  substr(description,1,300),topics,url
  FROM repos ORDER BY bucket, stars DESC""")
add_sheet("Repos",
  ["bucket","owner","name","role","category","visibility","fork","archived","language",
   "stars","forks","last_push","created","description","topics","url"], rows,
  [14,16,30,18,26,9,6,8,14,7,7,12,12,50,30,45])

# ---------- 2. Forks learning catalog ----------
rows = q("""SELECT full_name,bucket,category,upstream,upstream_stars,upstream_lang,
  upstream_age_years,staleness_years,
  CASE WHEN upstream_archived THEN 'ARCHIVED' ELSE '' END,
  substr(upstream_desc,1,300),upstream_url,our_url,learn_notes
  FROM forks_catalog ORDER BY upstream_stars DESC""")
add_sheet("Forks_Catalog",
  ["our_fork","bucket","theme","upstream","up_stars","up_lang","up_age_yrs",
   "our_fork_age_yrs","up_status","upstream_description","upstream_url","our_url","learn_notes"],
  rows, [34,13,24,34,9,12,10,12,10,55,42,42,40])

# ---------- 3. Accounts (counter-intel) ----------
rows = q("""SELECT login,relation,type,name,
  CASE WHEN security_signal THEN 'SEC' ELSE '' END,
  interests,company,location,followers_n,following_n,public_repos,
  substr(bio,1,200),website,twitter,substr(created_at,1,10),threat_notes,url
  FROM accounts WHERE relation IN ('follower','following','mutual')
  ORDER BY (security_signal=1) DESC, followers_n DESC""")
add_sheet("Accounts_1hop",
  ["login","relation","type","name","sec","interests","company","location",
   "followers","following","repos","bio","website","twitter","joined","notes","url"],
  rows, [22,10,8,24,5,26,20,20,9,9,7,45,30,16,12,28,40])

# ---------- 4. Their repos ----------
rows = q("""SELECT owner,full_name,
  CASE WHEN is_fork THEN 'fork' ELSE '' END,language,stars,substr(pushed_at,1,10),
  substr(description,1,250),url FROM account_repos ORDER BY stars DESC""")
add_sheet("Account_Repos",
  ["owner","repo","fork","language","stars","last_push","description","url"], rows,
  [22,40,6,14,7,12,55,45])

# ---------- 5. Threat model: 2-hop connectors (who is most connected to your network) ----------
rows = q("""
  SELECT e.dst AS node, count(*) AS hops_in,
    COALESCE(a.relation,'2-hop') rel,
    COALESCE(a.security_signal,0),
    substr(a.bio,1,160), a.followers_n, a.url
  FROM edges e LEFT JOIN accounts a ON a.login=e.dst
  WHERE e.src != ?
  GROUP BY e.dst HAVING hops_in >= 2
  ORDER BY hops_in DESC LIMIT 1000""", (ME,))
add_sheet("ThreatModel_Connectors",
  ["node","incoming_links_from_your_net","relation","sec_signal","bio","followers","url"],
  [(r[0],r[1],r[2],('SEC' if r[3] else ''),r[4],r[5],r[6]) for r in rows],
  [24,16,12,10,55,10,40])

# ---------- 6. Security-flagged accounts ----------
rows = q("""SELECT login,relation,name,interests,company,location,followers_n,
  substr(bio,1,220),url FROM accounts
  WHERE security_signal=1 ORDER BY relation, followers_n DESC""")
add_sheet("Security_Flagged",
  ["login","relation","name","interests","company","location","followers","bio","url"], rows,
  [22,10,24,26,20,20,9,55,40])

# ---------- 7. SANITIZATION CANDIDATES (review-only; nothing is deleted/unfollowed) ----------
# 7a. PRIORITY: unfollow-noise. Accounts you follow that look low-value:
#     inactive (>2yr), not mutual, no shared interest with your work, or mass-follow bots.
rows = q("""
  SELECT login, name, COALESCE(public_repos,0), COALESCE(followers_n,0), COALESCE(following_n,0),
    substr(COALESCE(interests,''),1,40), substr(updated_at,1,10),
    TRIM(
      CASE WHEN COALESCE(public_repos,0)=0 THEN 'no-repos ' ELSE '' END ||
      CASE WHEN updated_at IS NOT NULL AND updated_at < '2024-06-18' THEN 'inactive>2yr ' ELSE '' END ||
      CASE WHEN COALESCE(following_n,0) > 1500 THEN 'mass-follower ' ELSE '' END ||
      CASE WHEN interests IS NULL OR interests='Other/Uncategorized' THEN 'no-shared-interest ' ELSE '' END
    ) AS reasons,
    url
  FROM accounts
  WHERE relation='following'           -- following-only (excludes mutuals)
  ORDER BY (CASE WHEN COALESCE(public_repos,0)=0 THEN 1 ELSE 0 END
           + CASE WHEN updated_at < '2024-06-18' THEN 1 ELSE 0 END
           + CASE WHEN COALESCE(following_n,0) > 1500 THEN 1 ELSE 0 END
           + CASE WHEN interests IS NULL OR interests='Other/Uncategorized' THEN 1 ELSE 0 END) DESC,
           followers_n ASC""")
add_sheet("Unfollow_Candidates",
  ["login","name","repos","followers","following","interests","last_active","why_noise","url"],
  [r for r in rows], [22,24,7,9,9,40,12,40,40])

# 7b. Dead-fork candidates: upstream archived OR stale >3yr AND our fork also untouched >2yr.
rows = q("""SELECT full_name, upstream,
    CASE WHEN upstream_archived THEN 'archived' ELSE printf('%.1fyr stale', upstream_age_years) END,
    staleness_years, category, upstream_url
  FROM forks_catalog
  WHERE (upstream_archived=1 OR upstream_age_years>3) AND staleness_years>2
  ORDER BY upstream_archived DESC, upstream_age_years DESC""")
add_sheet("DeadFork_Candidates",
  ["our_fork","upstream","upstream_status","our_fork_age_yrs","theme","upstream_url"], rows,
  [34,34,14,14,24,42])

# 7c. Risky accounts: security-signal or thin/suspicious (new + no repos) among followers/2-hop.
rows = q("""SELECT login, relation,
    CASE WHEN security_signal THEN 'SEC' ELSE '' END,
    COALESCE(public_repos,0), COALESCE(followers_n,0), substr(created_at,1,10),
    TRIM(CASE WHEN security_signal THEN 'security-signal ' ELSE '' END ||
         CASE WHEN COALESCE(public_repos,0)=0 AND created_at>'2024-06-18' THEN 'new+empty ' ELSE '' END ||
         CASE WHEN COALESCE(following_n,0)>2000 AND COALESCE(followers_n,0)<10 THEN 'follow-bot ' ELSE '' END),
    substr(bio,1,160), url
  FROM accounts
  WHERE (security_signal=1 OR (public_repos=0 AND created_at>'2024-06-18')
         OR (following_n>2000 AND followers_n<10))
    AND relation IN ('follower','mutual','2-hop')
  ORDER BY (security_signal=1) DESC, relation""")
add_sheet("Risky_Accounts",
  ["login","relation","sec","repos","followers","joined","flags","bio","url"], rows,
  [22,10,5,7,9,12,28,55,40])

# 7d. Visibility audit: your PUBLIC org-originals (exposed surface near your network).
rows = q("""SELECT bucket, full_name, category, stars, substr(pushed_at,1,10),
    substr(description,1,200), url
  FROM repos WHERE role='org-original' AND private=0
  ORDER BY bucket, pushed_at DESC""")
add_sheet("Visibility_Audit",
  ["bucket","public_org_original","category","stars","last_push","description","url"], rows,
  [14,40,26,7,12,55,42])

# ---------- 8. Summary ----------
def scalar(sql,p=()): return db.execute(sql,p).fetchone()[0]
S = []
S.append(["=== REPOS ===",""])
S.append(["total repos", scalar("SELECT count(*) FROM repos")])
for r in q("SELECT bucket||' / '||role, count(*) FROM repos GROUP BY bucket,role ORDER BY bucket"):
    S.append([r[0], r[1]])
S.append(["",""]); S.append(["=== FORKS CATALOG ===",""])
S.append(["forks catalogued", scalar("SELECT count(*) FROM forks_catalog")])
S.append(["upstream resolved", scalar("SELECT count(*) FROM forks_catalog WHERE upstream IS NOT NULL")])
S.append(["upstream archived (dead projects)", scalar("SELECT count(*) FROM forks_catalog WHERE upstream_archived=1")])
S.append(["upstream stale >3yr", scalar("SELECT count(*) FROM forks_catalog WHERE upstream_age_years>3")])
S.append(["",""]); S.append(["=== SOCIAL / COUNTER-INTEL ===",""])
S.append(["1-hop accounts", scalar("SELECT count(*) FROM accounts WHERE relation IN ('follower','following','mutual')")])
S.append(["  of which crawled", scalar("SELECT count(*) FROM accounts WHERE crawled=1 AND relation IN ('follower','following','mutual')")])
S.append(["mutual", scalar("SELECT count(*) FROM accounts WHERE relation='mutual'")])
S.append(["followers-only", scalar("SELECT count(*) FROM accounts WHERE relation='follower'")])
S.append(["following-only", scalar("SELECT count(*) FROM accounts WHERE relation='following'")])
S.append(["2-hop accounts discovered", scalar("SELECT count(*) FROM accounts WHERE relation='2-hop'")])
S.append(["total follow-edges", scalar("SELECT count(*) FROM edges")])
S.append(["security-signal accounts (1-hop)", scalar("SELECT count(*) FROM accounts WHERE security_signal=1 AND relation IN ('follower','following','mutual')")])
ws = wb.create_sheet("Summary")
for c,h in enumerate(["metric","count"],1):
    cell=ws.cell(1,c); cell.value=h; cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2F5496")
for row in S: ws.append(row)
ws.column_dimensions["A"].width=40; ws.column_dimensions["B"].width=14
wb.move_sheet("Summary", -(len(wb.sheetnames)-1))  # Summary first

del wb["Sheet"]
out = os.path.join(WS, "github_inventory.xlsx")
wb.save(out)
print("WROTE", out)
for s in wb.sheetnames:
    print(f"  - {s}: {wb[s].max_row-1} rows")
