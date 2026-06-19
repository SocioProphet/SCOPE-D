#!/usr/bin/env python3
"""Mine repos of security-flagged SG network accounts for TTP inventory.

Pulls public repos for each security-flagged + crawled SG account, categorizes
them, detects offensive/defensive tooling patterns, and builds a TTP map.

Output:
  - raw/sg_security_repos.jsonl  — one record per repo
  - sg_ttp_inventory.xlsx        — TTP map with account → tools → categories
"""
import json, os, re, sqlite3, subprocess, time
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

WS = Path(os.path.expanduser("~/dev/gh-inventory"))
db = sqlite3.connect(WS / "inventory.db")
db.row_factory = sqlite3.Row

OUT_JSONL = WS / "raw" / "sg_security_repos.jsonl"
OUT_XLSX  = WS / "sg_ttp_inventory.xlsx"

OFFENSIVE_PATTERNS = [
    (re.compile(r"exploit|payload|shellcode|rop.chain|ret2", re.I),             "exploit_dev"),
    (re.compile(r"c2|command.and.control|beacon|implant|rat\b|trojan", re.I),   "c2_implant"),
    (re.compile(r"phish|credential.harvest|spearphish", re.I),                  "phishing"),
    (re.compile(r"recon|osint|footprint|enumerat", re.I),                       "recon_osint"),
    (re.compile(r"fuzzer|fuzz|afl|honggfuzz|libfuzzer", re.I),                  "fuzzing"),
    (re.compile(r"bypass|evasion|av.bypass|amsi|edr.bypass", re.I),             "evasion"),
    (re.compile(r"privesc|privilege.escal|lpe\b|sudo.exploit", re.I),           "privesc"),
    (re.compile(r"lateral.move|pass.the.hash|kerberoast|bloodhound", re.I),     "lateral_movement"),
    (re.compile(r"exfil|data.theft|exfiltrat", re.I),                           "exfiltration"),
    (re.compile(r"ransomware|encryptor|locker", re.I),                          "ransomware"),
    (re.compile(r"rootkit|bootkit|kernel.exploit|ring0", re.I),                 "rootkit"),
    (re.compile(r"injection|sqli|xss|rce\b|lfi\b|rfi\b|ssrf\b|idor\b", re.I), "web_vuln"),
    (re.compile(r"scanner|port.scan|nmap|masscan|network.scan", re.I),          "network_scan"),
    (re.compile(r"password|crack|hashcat|john.the.ripper|brute.force", re.I),   "credential_attack"),
    (re.compile(r"malware|virus|worm|botnet", re.I),                            "malware"),
    (re.compile(r"mcp.poison|tool.poison|prompt.inject|jailbreak", re.I),       "ai_attack"),
]

DEFENSIVE_PATTERNS = [
    (re.compile(r"detect|sigma|yara|snort|suricata|splunk", re.I),      "detection"),
    (re.compile(r"harden|cis.bench|stig|compliance|baseline", re.I),    "hardening"),
    (re.compile(r"honeypot|canary|deception|tarpit", re.I),             "deception"),
    (re.compile(r"siem|log.analytic|elk|opensearch", re.I),             "siem"),
    (re.compile(r"threat.intel|cti|ioc|stix|taxii", re.I),              "threat_intel"),
    (re.compile(r"incident.response|ir\b|forensic|dfir", re.I),         "dfir"),
    (re.compile(r"pentest|red.team|purple.team|ctf\b", re.I),           "pentest"),
    (re.compile(r"vuln.mgmt|patch|cve|nvd|cvss", re.I),                 "vuln_mgmt"),
]

AI_PATTERNS = [
    (re.compile(r"llm|gpt|claude|openai|anthropic|gemini|mistral", re.I), "llm_tools"),
    (re.compile(r"agent|agentic|mcp|tool.call|function.call", re.I),     "agent_framework"),
    (re.compile(r"prompt|system.prompt|jailbreak|bypass", re.I),         "prompt_engineering"),
    (re.compile(r"embedding|vector|rag|retrieval", re.I),                "rag_vector"),
    (re.compile(r"fine.tun|lora|qlora|training|finetune", re.I),        "model_training"),
]

def classify(name, desc, topics, lang):
    text = f"{name} {desc or ''} {topics or ''} {lang or ''}".lower()
    off = [t for p, t in OFFENSIVE_PATTERNS if p.search(text)]
    defs = [t for p, t in DEFENSIVE_PATTERNS if p.search(text)]
    ai  = [t for p, t in AI_PATTERNS if p.search(text)]
    return off, defs, ai

def gh_repos(login):
    try:
        r = subprocess.run(
            ["gh", "api", f"users/{login}/repos", "--paginate",
             "--jq", "[.[] | {name:.name, desc:.description, lang:.language, topics:(.topics|join(\",\")), stars:.stargazers_count, fork:.fork, url:.html_url}]"],
            capture_output=True, text=True, timeout=20
        )
        if r.returncode != 0:
            return []
        items = []
        for line in r.stdout.strip().splitlines():
            try:
                batch = json.loads(line)
                items.extend(batch if isinstance(batch, list) else [batch])
            except Exception:
                pass
        return [x for x in items if not x.get("fork")]
    except Exception:
        return []

accounts = db.execute("""
    SELECT login, bio, followers_n, following_n, interests, relation, url
    FROM sg_accounts
    WHERE security_signal=1 AND crawled=1
    ORDER BY followers_n DESC
""").fetchall()

print(f"Mining repos for {len(accounts)} security-flagged SG accounts...")

OUT_JSONL.parent.mkdir(parents=True, exist_ok=True)
seen = set()
if OUT_JSONL.exists():
    for line in OUT_JSONL.read_text().splitlines():
        try:
            seen.add(json.loads(line)["login"])
        except Exception:
            pass

ttp_map   = defaultdict(lambda: {"off":set(), "def":set(), "ai":set(), "repos":[], "stars":0})
all_repos = []

for i, acct in enumerate(accounts):
    login = acct["login"]
    if login in seen:
        continue
    repos = gh_repos(login)
    if not repos:
        time.sleep(0.2)
        continue

    record = {
        "login":    login,
        "bio":      acct["bio"],
        "followers": acct["followers_n"],
        "relation": acct["relation"],
        "url":      acct["url"],
        "repos":    repos,
        "repo_count": len(repos),
        "offensive_ttps": [],
        "defensive_ttps": [],
        "ai_ttps": [],
    }
    for repo in repos:
        off, defs, ai = classify(repo["name"], repo.get("desc"), repo.get("topics"), repo.get("lang"))
        record["offensive_ttps"].extend(off)
        record["defensive_ttps"].extend(defs)
        record["ai_ttps"].extend(ai)
        ttp_map[login]["off"].update(off)
        ttp_map[login]["def"].update(defs)
        ttp_map[login]["ai"].update(ai)
        ttp_map[login]["repos"].append(repo["name"])
        ttp_map[login]["stars"] += repo.get("stars", 0)

    record["offensive_ttps"] = list(set(record["offensive_ttps"]))
    record["defensive_ttps"] = list(set(record["defensive_ttps"]))
    record["ai_ttps"]        = list(set(record["ai_ttps"]))

    with open(OUT_JSONL, "a") as f:
        f.write(json.dumps(record) + "\n")
    seen.add(login)
    all_repos.extend(repos)

    if i % 25 == 0:
        print(f"  [{i+1}/{len(accounts)}] {login}: {len(repos)} repos, off={record['offensive_ttps'][:3]}")

    time.sleep(0.3)

print(f"\nMined {len(seen)} accounts, {len(all_repos)} repos total")

# ── TTP aggregate stats ────────────────────────────────────────────────────────
off_agg  = defaultdict(int)
def_agg  = defaultdict(int)
ai_agg   = defaultdict(int)
for login, data in ttp_map.items():
    for t in data["off"]:  off_agg[t]  += 1
    for t in data["def"]:  def_agg[t]  += 1
    for t in data["ai"]:   ai_agg[t]   += 1

print("\nOffensive TTP prevalence (accounts with this TTP):")
for t, n in sorted(off_agg.items(), key=lambda x: -x[1]):
    print(f"  {t:25} {n}")
print("\nDefensive TTP prevalence:")
for t, n in sorted(def_agg.items(), key=lambda x: -x[1]):
    print(f"  {t:25} {n}")
print("\nAI TTP prevalence:")
for t, n in sorted(ai_agg.items(), key=lambda x: -x[1]):
    print(f"  {t:25} {n}")

# ── xlsx ───────────────────────────────────────────────────────────────────────
HDR = PatternFill("solid", fgColor="1F3864")
OFF = PatternFill("solid", fgColor="C00000")
DEF = PatternFill("solid", fgColor="375623")

wb = Workbook()
ws = wb.active
ws.title = "SG_TTP_Map"

def make_sheet(wb, title, headers, rows, widths, key_col=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    if rows: ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

account_rows = []
for login, data in sorted(ttp_map.items(), key=lambda x: -(len(x[1]["off"])*3 + len(x[1]["def"]))):
    account_rows.append([
        login,
        f"https://github.com/{login}",
        data["stars"],
        len(data["repos"]),
        ", ".join(sorted(data["off"])) or "—",
        ", ".join(sorted(data["def"])) or "—",
        ", ".join(sorted(data["ai"])) or "—",
        "HIGH" if data["off"] else ("MEDIUM" if data["def"] else "LOW"),
    ])

del wb["Sheet"]
make_sheet(wb, "SG_TTP_Map",
    ["login","url","total_stars","repo_count","offensive_ttps","defensive_ttps","ai_ttps","threat_level"],
    account_rows, [24,42,10,10,55,45,40,12])

make_sheet(wb, "TTP_Prevalence",
    ["ttp","type","accounts_with_ttp"],
    sorted(
        [(t,"offensive",n) for t,n in off_agg.items()] +
        [(t,"defensive",n) for t,n in def_agg.items()] +
        [(t,"ai",n) for t,n in ai_agg.items()],
        key=lambda x: -x[2]
    ), [28, 12, 20])

wb.save(OUT_XLSX)
print(f"\nWrote {OUT_XLSX}")
db.close()
